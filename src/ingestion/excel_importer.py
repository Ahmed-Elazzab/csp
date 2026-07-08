"""
Excel ingestion module.

Reads 'Attributes' and 'Scenario Questions' sheets from the workbook and
seeds the database reference tables. Safe to run multiple times (upserts).
"""

from __future__ import annotations

import logging
import os
from typing import Optional

import openpyxl

from src.config import get_settings
from src.database.connection import get_db_session
from src.database.models import AttributeDefinition, QuestionnaireQuestion

logger = logging.getLogger(__name__)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _str(val: object) -> Optional[str]:
    """Convert a cell value to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _bool(val: object) -> bool:
    """Interpret TRUE/FALSE strings or booleans."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() == "TRUE"
    return False


# ── Public functions ──────────────────────────────────────────────────────────

def load_attributes_from_excel(path: str) -> list[dict]:
    """Parse 'Attributes' sheet and return list of attribute dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Attributes"]

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]

    records = []
    for row in rows[1:]:
        if not any(row):
            continue
        data = dict(zip(header, row))
        records.append(
            {
                "name": _str(data.get("Attribute Name")),
                "category": _str(data.get("Category")) or "Uncategorised",
                "impact": _str(data.get("Impact on Performance")),
                "reasons": _str(data.get("Reasons")),
                "data_type": _str(data.get("Data Type")),
                "sample_values": _str(data.get("Sample Values")),
                "attribute_source": _str(data.get("Attribute Source")),
                "currently_exists": _bool(data.get("Currently Exists")),
            }
        )
    wb.close()
    return [r for r in records if r["name"]]


def load_questions_from_excel(path: str) -> list[dict]:
    """Parse 'Scenario Questions' sheet and return list of question dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Scenario Questions"]

    rows = list(ws.iter_rows(values_only=True))
    # Skip header row
    records = []
    for row in rows[1:]:
        if not any(row):
            continue
        scenario, qid, question = (row + (None, None, None))[:3]
        if scenario and qid and question:
            records.append(
                {
                    "scenario": _str(scenario),
                    "question_id": _str(qid),
                    "question_text": _str(question),
                }
            )
    wb.close()
    return records


def seed_attributes(path: str) -> int:
    """Upsert attribute definitions. Returns count of records processed."""
    records = load_attributes_from_excel(path)
    inserted = 0
    with get_db_session() as session:
        for rec in records:
            existing = (
                session.query(AttributeDefinition)
                .filter_by(name=rec["name"])
                .first()
            )
            if existing:
                for k, v in rec.items():
                    if k != "name":
                        setattr(existing, k, v)
            else:
                session.add(AttributeDefinition(**rec))
                inserted += 1
    logger.info("Attributes seeded: %d processed, %d new", len(records), inserted)
    return len(records)


def seed_questions(path: str) -> int:
    """Upsert questionnaire questions. Returns count of records processed."""
    records = load_questions_from_excel(path)
    inserted = 0
    with get_db_session() as session:
        for rec in records:
            existing = (
                session.query(QuestionnaireQuestion)
                .filter_by(scenario=rec["scenario"], question_id=rec["question_id"])
                .first()
            )
            if existing:
                existing.question_text = rec["question_text"]
            else:
                session.add(QuestionnaireQuestion(**rec))
                inserted += 1
    logger.info("Questions seeded: %d processed, %d new", len(records), inserted)
    return len(records)


def seed_database_if_empty() -> None:
    """
    Seed attribute definitions and questionnaire questions if the tables are empty.
    Called once at application startup.
    """
    settings = get_settings()
    excel_path = settings.EXCEL_PATH

    if not os.path.exists(excel_path):
        logger.warning("Excel file not found at %s – skipping seed", excel_path)
        return

    with get_db_session() as session:
        attr_count = session.query(AttributeDefinition).count()
        q_count = session.query(QuestionnaireQuestion).count()

    if attr_count == 0:
        seed_attributes(excel_path)
    else:
        logger.info("Attribute definitions already seeded (%d rows)", attr_count)

    if q_count == 0:
        seed_questions(excel_path)
    else:
        logger.info("Questionnaire questions already seeded (%d rows)", q_count)


def force_reseed() -> tuple[int, int]:
    """Re-import attributes and questions from Excel, overwriting existing rows."""
    settings = get_settings()
    path = settings.EXCEL_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel not found: {path}")
    a = seed_attributes(path)
    q = seed_questions(path)
    return a, q
